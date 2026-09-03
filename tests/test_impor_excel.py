"""Tes parser impor: deteksi header, kolom tipologi, dan normalisasi nama."""
import sys
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402

from services import impor_excel, normalisasi  # noqa: E402


def _sheet_contoh(baris_header: int = 3, kolom_tipologi=("T1", "T2", "T3")):
    """Sheet tiruan dengan header di posisi yang bisa digeser-geser."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contoh"
    kepala = ["No", "Nama Objek", "DESA", "KECAMATAN", "KABUPATEN", "Nama Wakif",
              "AIW/APAIW", "LUAS PERSIL", "TIPOLOGI PERMASALAHAN"]
    for i, nama in enumerate(kepala, 2):
        ws.cell(baris_header, i, nama)
    kolom_awal = 2 + len(kepala) - 1
    for i, kode in enumerate(kolom_tipologi):
        ws.cell(baris_header + 1, kolom_awal + i, kode)
    return wb, ws, baris_header, kolom_awal


class TesDeteksiHeader(unittest.TestCase):
    def test_header_di_baris_tiga(self):
        _, ws, hdr, _ = _sheet_contoh(baris_header=3)
        self.assertEqual(impor_excel.cari_baris_header(ws), 3)

    def test_header_di_baris_satu(self):
        _, ws, hdr, _ = _sheet_contoh(baris_header=1)
        self.assertEqual(impor_excel.cari_baris_header(ws), 1)

    def test_kolom_dipetakan_berdasarkan_nama_bukan_posisi(self):
        _, ws, hdr, _ = _sheet_contoh(baris_header=3)
        peta = impor_excel.petakan_kolom(ws, hdr)
        self.assertEqual(peta["nama_objek"], 3)
        self.assertEqual(peta["luas_persil"], 9)
        # Geser satu kolom ke kanan: nomor kolomnya ikut bergeser.
        wb2 = Workbook()
        ws2 = wb2.active
        for i, nama in enumerate(["No", "Nama Objek", "DESA"], 5):
            ws2.cell(3, i, nama)
        peta2 = impor_excel.petakan_kolom(ws2, 3)
        self.assertEqual(peta2["nama_objek"], 6)

    def test_kolom_kkp_tidak_tertelan_kolom_kecamatan(self):
        """Header 'KECAMATAN KKP' harus terpetakan sendiri, bukan jadi 'kecamatan'."""
        from openpyxl import Workbook as WB
        wb = WB()
        ws = wb.active
        for i, nama in enumerate(
                ["Nama Objek", "KECAMATAN", "KECAMATAN  KKP", "KELURAHAN/DESA KKP"], 1):
            ws.cell(3, i, nama)
        peta = impor_excel.petakan_kolom(ws, 3)
        self.assertEqual(peta["kecamatan"], 2)
        self.assertEqual(peta["kecamatan_kkp"], 3)
        self.assertEqual(peta["desa_kkp"], 4)

    def test_sheet_tanpa_header_dilaporkan(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1, "tidak ada apa-apa")
        laporan = impor_excel.Laporan()
        self.assertEqual(impor_excel.baca_sheet(ws, laporan), [])
        self.assertTrue(laporan.peringatan)


class TesKolomTipologi(unittest.TestCase):
    def test_tujuh_kolom_terpetakan(self):
        _, ws, hdr, awal = _sheet_contoh(
            baris_header=3, kolom_tipologi=("T1", "T2", "T3", "T4", "T5", "T6", "T7"))
        peta = impor_excel.petakan_tipologi(ws, hdr + 1)
        self.assertEqual(len(peta), 7)
        self.assertEqual(peta["T1"], awal)
        self.assertEqual(peta["T7"], awal + 6)

    def test_sheet_hanya_sampai_t6(self):
        """Tujuh sheet di file asli tidak punya kolom T7 sama sekali."""
        _, ws, hdr, _ = _sheet_contoh(
            baris_header=3, kolom_tipologi=("T1", "T2", "T3", "T4", "T5", "T6"))
        peta = impor_excel.petakan_tipologi(ws, hdr + 1)
        self.assertNotIn("T7", peta)
        self.assertEqual(len(peta), 6)

    def test_centang_tunggal_jadi_satu_kode(self):
        wb, ws, hdr, awal = _sheet_contoh(
            baris_header=3, kolom_tipologi=("T1", "T2", "T3"))
        ws.cell(5, 2, 1)
        ws.cell(5, 3, "Masjid Contoh")
        ws.cell(5, awal + 1, "√")
        laporan = impor_excel.Laporan()
        hasil = impor_excel.baca_sheet(ws, laporan)
        self.assertEqual(len(hasil), 1)
        self.assertEqual(hasil[0]["tipologi_kode"], "T2")

    def test_tanpa_centang_jadi_none(self):
        wb, ws, hdr, awal = _sheet_contoh(baris_header=3)
        ws.cell(5, 2, 1)
        ws.cell(5, 3, "Masjid Kosong")
        hasil = impor_excel.baca_sheet(ws, impor_excel.Laporan())
        self.assertIsNone(hasil[0]["tipologi_kode"])

    def test_centang_ganda_ambil_kompleksitas_tertinggi(self):
        self.assertEqual(impor_excel.pilih_tipologi(["T2", "T6"]), "T6")
        self.assertEqual(impor_excel.pilih_tipologi(["T7", "T1"]), "T7")
        self.assertIsNone(impor_excel.pilih_tipologi([]))

    def test_centang_ganda_dicatat_di_peringatan(self):
        wb, ws, hdr, awal = _sheet_contoh(baris_header=3)
        ws.cell(5, 2, 1)
        ws.cell(5, 3, "Masjid Ganda")
        ws.cell(5, awal, "√")
        ws.cell(5, awal + 2, "√")
        laporan = impor_excel.Laporan()
        hasil = impor_excel.baca_sheet(ws, laporan)
        self.assertEqual(hasil[0]["tipologi_kode"], "T3")
        self.assertEqual(len(laporan.peringatan), 1)


class TesBarisBukanData(unittest.TestCase):
    def test_baris_tanpa_nomor_urut_dilewati(self):
        """Blok legenda di bawah data ikut mengisi kolom Nama Objek."""
        wb, ws, hdr, _ = _sheet_contoh(baris_header=3)
        ws.cell(5, 2, 1)
        ws.cell(5, 3, "Masjid Asli")
        ws.cell(9, 3, "Tidak bisa ditindaklanjut :")
        laporan = impor_excel.Laporan()
        hasil = impor_excel.baca_sheet(ws, laporan)
        self.assertEqual([b["nama_objek"] for b in hasil], ["Masjid Asli"])
        self.assertEqual(len(laporan.bukan_data), 1)


class TesNormalisasiNama(unittest.TestCase):
    def test_ejaan_kecamatan_disatukan(self):
        self.assertEqual(normalisasi.kecamatan("Bonepantai"), "Bone Pantai")
        self.assertEqual(normalisasi.kecamatan("bone pantai"), "Bone Pantai")
        self.assertEqual(normalisasi.kecamatan("  Bone   Pantai "), "Bone Pantai")

    def test_ejaan_kabupaten_disatukan(self):
        for teks in ("Bone Bolango", "Kabupaten Bone bolango", "Kabupaten Bone Bolango"):
            self.assertEqual(normalisasi.kabupaten(teks), "Kabupaten Bone Bolango")

    def test_spasi_ganda_desa_dirapikan(self):
        self.assertEqual(normalisasi.desa("Bulontala  Timur"), "Bulontala Timur")
        self.assertIsNone(normalisasi.desa("  "))
        self.assertIsNone(normalisasi.desa("-"))

    def test_luas_tipe_campur(self):
        self.assertEqual(normalisasi.angka("264"), 264.0)
        self.assertEqual(normalisasi.angka(410.0), 410.0)
        self.assertEqual(normalisasi.angka("1.234,5"), 1234.5)
        self.assertEqual(normalisasi.angka("340 m²"), 340.0)
        self.assertIsNone(normalisasi.angka(""))
        self.assertIsNone(normalisasi.angka(None))

    def test_koordinat_hanya_dari_url_yang_masuk_akal(self):
        lat, lon = normalisasi.koordinat_dari_url(
            "https://maps.google.com/?q=0.567890,123.456789")
        self.assertAlmostEqual(lat, 0.56789)
        self.assertAlmostEqual(lon, 123.456789)
        self.assertEqual(normalisasi.koordinat_dari_url("511079.0"), (None, None))
        self.assertEqual(normalisasi.koordinat_dari_url(None), (None, None))

    def test_centang(self):
        self.assertTrue(normalisasi.dicentang("√"))
        self.assertTrue(normalisasi.dicentang(" v "))
        self.assertFalse(normalisasi.dicentang(""))
        self.assertFalse(normalisasi.dicentang(None))


if __name__ == "__main__":
    unittest.main()
