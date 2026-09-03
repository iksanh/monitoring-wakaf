"""Ekspor rekap dan daftar objek ke .xlsx (openpyxl)."""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config

_FONT = "Arial"
_ISI = Font(name=_FONT, size=10)
_JUDUL = Font(name=_FONT, size=12, bold=True)
_KEPALA = Font(name=_FONT, size=10, bold=True, color="FFFFFF")
_LATAR_KEPALA = PatternFill("solid", start_color="0B6B3A")


def _tulis(ws, judul: str, kolom: list[str], baris: list[list]) -> None:
    ws.cell(1, 1, judul).font = _JUDUL
    ws.cell(2, 1, config.NAMA_KANTOR).font = _ISI
    for i, nama in enumerate(kolom, 1):
        sel = ws.cell(4, i, nama)
        sel.font = _KEPALA
        sel.fill = _LATAR_KEPALA
        sel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r, isi in enumerate(baris, 5):
        for c, nilai in enumerate(isi, 1):
            sel = ws.cell(r, c, nilai)
            sel.font = _ISI
            sel.alignment = Alignment(vertical="top", wrap_text=isinstance(nilai, str))
    _autofit(ws, len(kolom))
    ws.freeze_panes = "A5"


def _autofit(ws, jumlah_kolom: int, maks: int = 45) -> None:
    for c in range(1, jumlah_kolom + 1):
        lebar = 8
        for sel in ws[get_column_letter(c)]:
            if sel.row < 4 or sel.value is None:
                continue
            lebar = max(lebar, min(maks, len(str(sel.value)) + 2))
        ws.column_dimensions[get_column_letter(c)].width = lebar


def _byte(wb) -> bytes:
    penampung = io.BytesIO()
    wb.save(penampung)
    return penampung.getvalue()


def daftar_objek(baris: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Objek Wakaf"
    kolom = ["Kode", "Nama Objek", "Desa", "Kecamatan", "Wilayah", "Wakif", "Nadzir",
             "AIW/APAIW", "Tipe Hak", "NIB", "Luas (m2)", "Tipologi", "Status Sertipikat",
             "Keterangan"]
    isi = [[b["kode"], b["nama_objek"], b["desa_nama"], b["kecamatan_nama"],
            b["wilayah_nama"], b["nama_wakif"], b["nama_nadzir"], b["no_aiw"],
            b["tipe_hak"], b["nib"], b["luas_persil"], b["tipologi_kode"],
            b["status_sertipikat"], b["keterangan"]] for b in baris]
    _tulis(ws, "DAFTAR OBJEK TANAH WAKAF", kolom, isi)
    return _byte(wb)


def rekap_harian(rekap: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Harian"
    kolom = ["Tahapan", "Aksi", "Objek Wakaf", "Kode", "No Berkas", "Kecamatan",
             "Wilayah", "Petugas", "Catatan"]
    isi = [[b["tahapan_nama"], b["aksi"], b["nama_objek"], b["objek_kode"],
            b["no_berkas"], b["kecamatan_nama"], b["wilayah_nama"],
            b["nama_pengguna"], b["catatan"]] for b in rekap["berkas"]]
    _tulis(ws, f"REKAP HARIAN PERGERAKAN BERKAS — {rekap['tanggal']}", kolom, isi)

    ws2 = wb.create_sheet("Ringkasan")
    _tulis(ws2, "RINGKASAN PER TAHAPAN", ["Tahapan", "Aksi", "Jumlah"],
           [[b["nama"], b["aksi"], b["jumlah"]] for b in rekap["per_tahapan"]])
    return _byte(wb)


def rekap_tahapan(rekap: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Per Tahapan"
    _tulis(ws, "POSISI BERKAS PER TAHAPAN", ["Tahapan", "Berkas Aktif", "Total Berkas"],
           [[b["nama"], b["aktif"], b["semua"]] for b in rekap["corong"]])
    return _byte(wb)


def rekap_potensi(baris: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Potensi"
    isi = [[b["kecamatan"], b["wilayah"], b["baru"], b["ada_hak"], b["isbat"], b["total"]]
           for b in baris]
    isi.append(["TOTAL", "", sum(b["baru"] or 0 for b in baris),
                sum(b["ada_hak"] or 0 for b in baris),
                sum(b["isbat"] or 0 for b in baris),
                sum(b["total"] or 0 for b in baris)])
    _tulis(ws, "TOTAL POTENSI SERTIPIKAT WAKAF",
           ["Kecamatan", "Wilayah", "Baru", "Ada Hak", "Isbat Wakaf", "Total"], isi)
    return _byte(wb)


def nama_berkas(dasar: str, tanggal: str | None = None) -> str:
    return f"{dasar}_{tanggal or config.hari_ini_iso()}.xlsx"
