"""Parser sheet kecamatan pada Rekapan_Wakaf_Bone_bolango.xlsx.

Dipisah dua lapis:
  baca_*  -> hanya membaca & menormalisasi, tidak menyentuh database (mudah dites)
  simpan  -> menulis ke database, idempoten
"""
import re
from dataclasses import dataclass, field

import openpyxl

import db
from services import normalisasi

SHEET_BUKAN_KECAMATAN = {"POTENSI WAKAF", "Total Potensi Wilayah", "TIPOLOGI"}
KODE_TIPOLOGI = [f"T{i}" for i in range(1, 8)]
# Urutan prioritas kalau satu baris tercentang lebih dari satu tipologi.
PRIORITAS_TIPOLOGI = {"T7": 7, "T6": 6, "T5": 5, "T4": 4, "T3": 3, "T2": 2, "T1": 1}

# Nama kolom di sheet -> nama field. Kunci dibandingkan setelah dinormalisasi.
PETA_KOLOM = {
    "no": "no_urut",
    "nama objek": "nama_objek",
    "desa": "desa",
    "kecamatan": "kecamatan",
    "kabupaten": "kabupaten",
    "nama wakif": "nama_wakif",
    "nama nadzir": "nama_nadzir",
    "aiw/apaiw": "no_aiw",
    "tipehak": "tipe_hak",
    "tipe hak": "tipe_hak",
    "nib": "nib",
    "luas persil": "luas_persil",
    "kecamatan kkp": "kecamatan_kkp",
    "kelurahan/desa kkp": "desa_kkp",
    "keterangan": "keterangan",
    "rtrw": "rtrw",
    "rekomendasi isbat": "rekomendasi_isbat",
    "catatan kua": "catatan_kua",
    "file sertifikat wakaf": "url_dokumen",
    "titik lokasi": "url_maps",
    "update": "update_terakhir",
}

PETA_KEMENAG = {
    "nama": "nama",
    "kabupaten": "kabupaten",
    "kecamatan": "kecamatan",
    "kategori": "kategori",
    "tahun_berd": "tahun_berdiri",
    "latitude": "latitude",
    "longitude": "longitude",
    "id_kemenag": "id_kemenag",
    "sumber": "sumber",
    "keterangan": "keterangan",
    "tipe hak": "tipe_hak",
    "sertipikat": "status_sertipikat",
    "tahun": "tahun",
    "nib": "nib",
    "luastertul": "luas",
}

_ID_KEMENAG = re.compile(r"^\d{2}\.\d\.\d{2}\.\d{2}\.\d{2}\.\d{6}$")


@dataclass
class Laporan:
    terbaca: dict = field(default_factory=dict)
    tersimpan: dict = field(default_factory=dict)
    dilewati: list = field(default_factory=list)
    peringatan: list = field(default_factory=list)
    bukan_data: list = field(default_factory=list)
    kemenag_terbaca: dict = field(default_factory=dict)
    kemenag_tersimpan: int = 0
    desa_baru: int = 0
    tipologi: dict = field(default_factory=dict)

    @property
    def total_terbaca(self) -> int:
        return sum(self.terbaca.values())

    @property
    def total_tersimpan(self) -> int:
        return sum(self.tersimpan.values())


def _kunci(nilai) -> str:
    teks = normalisasi.rapikan(nilai)
    return (teks or "").lower().replace("(m²)", "").replace("(m2)", "").strip()


def _sel(ws, baris, kolom):
    return ws.cell(row=baris, column=kolom).value


def cari_baris_header(ws, maks_baris: int = 8) -> int | None:
    """Baris header dikenali dari sel berisi 'Nama Objek', bukan dari posisi tetap."""
    for r in range(1, min(maks_baris, ws.max_row or 1) + 1):
        for c in range(1, (ws.max_column or 1) + 1):
            if _kunci(_sel(ws, r, c)) == "nama objek":
                return r
    return None


def petakan_kolom(ws, baris_header: int) -> dict:
    """Peta nama field -> nomor kolom. Berdasarkan NAMA, bukan posisi.

    Kecocokan persis didahulukan, lalu prefiks TERPANJANG. Kalau dibalik,
    header 'KECAMATAN KKP' akan tertelan oleh pola 'kecamatan'.
    """
    peta = {}
    for c in range(1, (ws.max_column or 1) + 1):
        nama = _kunci(_sel(ws, baris_header, c))
        if not nama:
            continue
        if nama in PETA_KOLOM:
            peta.setdefault(PETA_KOLOM[nama], c)
            continue
        cocok = [p for p in PETA_KOLOM if nama.startswith(p)]
        if cocok:
            peta.setdefault(PETA_KOLOM[max(cocok, key=len)], c)
    return peta


def petakan_tipologi(ws, baris_sub: int) -> dict:
    """Kolom T1..T7 dibaca dari baris sub-header. Jumlah kolomnya beda antar sheet."""
    peta = {}
    for c in range(1, (ws.max_column or 1) + 1):
        nama = (normalisasi.rapikan(_sel(ws, baris_sub, c)) or "").upper()
        if nama in KODE_TIPOLOGI:
            peta.setdefault(nama, c)
    return peta


def pilih_tipologi(tercentang: list[str]) -> str | None:
    """Kalau lebih dari satu tercentang, ambil kompleksitas tertinggi."""
    if not tercentang:
        return None
    return max(tercentang, key=lambda k: PRIORITAS_TIPOLOGI.get(k, 0))


def baca_sheet(ws, laporan: Laporan) -> list[dict]:
    """Baca satu sheet kecamatan menjadi daftar dict objek wakaf."""
    baris_header = cari_baris_header(ws)
    if baris_header is None:
        laporan.peringatan.append(f"{ws.title}: baris header tidak ditemukan, sheet dilewati.")
        return []
    kolom = petakan_kolom(ws, baris_header)
    tipo = petakan_tipologi(ws, baris_header + 1)
    if "nama_objek" not in kolom:
        laporan.peringatan.append(f"{ws.title}: kolom 'Nama Objek' tidak terbaca.")
        return []

    hasil = []
    for r in range(baris_header + 2, (ws.max_row or 0) + 1):
        nama = normalisasi.rapikan(_sel(ws, r, kolom["nama_objek"]))
        if not nama:
            continue
        ambil = lambda f: _sel(ws, r, kolom[f]) if f in kolom else None  # noqa: E731

        # Di bawah data utama tiap sheet ada blok legenda dan rekap kecil yang
        # ikut mengisi kolom Nama Objek. Baris data asli selalu bernomor urut.
        if normalisasi.angka(ambil("no_urut")) is None:
            laporan.bukan_data.append(f"{ws.title} baris {r}: '{nama}' tanpa nomor urut.")
            continue

        tercentang = [k for k, c in tipo.items() if normalisasi.dicentang(_sel(ws, r, c))]
        if len(tercentang) > 1:
            laporan.peringatan.append(
                f"{ws.title} baris {r} ({nama}): tipologi ganda {sorted(tercentang)} "
                f"-> dipakai {pilih_tipologi(tercentang)}."
            )
        url_maps = normalisasi.url(ambil("url_maps"))
        lat, lon = normalisasi.koordinat_dari_url(url_maps)

        kec = normalisasi.kecamatan(ambil("kecamatan")) or normalisasi.kecamatan(ws.title)
        hasil.append({
            "sheet": ws.title,
            "baris": r,
            "nama_objek": nama,
            "desa": normalisasi.desa(ambil("desa")),
            "kecamatan": kec,
            "kabupaten": normalisasi.kabupaten(ambil("kabupaten")),
            "nama_wakif": normalisasi.rapikan(ambil("nama_wakif")),
            "nama_nadzir": normalisasi.rapikan(ambil("nama_nadzir")),
            "no_aiw": normalisasi.rapikan(ambil("no_aiw")),
            "tipe_hak": normalisasi.rapikan(ambil("tipe_hak")),
            "nib": normalisasi.rapikan(ambil("nib")),
            "luas_persil": normalisasi.angka(ambil("luas_persil")),
            "kecamatan_kkp": normalisasi.kecamatan(ambil("kecamatan_kkp")),
            "desa_kkp": normalisasi.desa(ambil("desa_kkp")),
            "keterangan": normalisasi.rapikan(ambil("keterangan")),
            "rtrw": normalisasi.rapikan(ambil("rtrw")),
            "tipologi_kode": pilih_tipologi(tercentang),
            "rekomendasi_isbat": normalisasi.rapikan(ambil("rekomendasi_isbat")),
            "catatan_kua": normalisasi.rapikan(ambil("catatan_kua")),
            "url_dokumen": normalisasi.url(ambil("url_dokumen")),
            "url_maps": url_maps,
            "latitude": lat,
            "longitude": lon,
        })
    laporan.terbaca[ws.title] = len(hasil)
    return hasil


def cari_blok_kemenag(ws) -> tuple[int, dict] | None:
    """Blok tempelan 'HASIL REKAPAN DATA GMAPS X KKP'.

    Header blok kadang bergeser satu kolom dari datanya (sheet Bone Pantai),
    jadi offset dihitung dari letak nyata nilai ID_KEMENAG di baris data pertama.
    """
    baris_header = kolom_id = None
    for r in range(1, min(10, ws.max_row or 1) + 1):
        for c in range(1, (ws.max_column or 1) + 1):
            if _kunci(_sel(ws, r, c)) == "id_kemenag":
                baris_header, kolom_id = r, c
                break
        if baris_header:
            break
    if not baris_header:
        return None

    geser = 0
    for r in range(baris_header + 1, min(baris_header + 6, (ws.max_row or 0)) + 1):
        for c in range(max(1, kolom_id - 3), min((ws.max_column or 1), kolom_id + 3) + 1):
            nilai = normalisasi.rapikan(_sel(ws, r, c))
            if nilai and _ID_KEMENAG.match(nilai):
                geser = c - kolom_id
                break
        if geser:
            break

    peta = {}
    for c in range(1, (ws.max_column or 1) + 1):
        nama = _kunci(_sel(ws, baris_header, c))
        if nama in PETA_KEMENAG and c >= min(kolom_id, kolom_id + geser) - 12:
            peta.setdefault(PETA_KEMENAG[nama], c + geser)
    if "kecamatan" in peta:
        peta.setdefault("desa", peta["kecamatan"] + 1)
    return (baris_header, peta) if "id_kemenag" in peta else None


def baca_kemenag(ws, laporan: Laporan) -> list[dict]:
    blok = cari_blok_kemenag(ws)
    if not blok:
        return []
    baris_header, peta = blok
    hasil = []
    for r in range(baris_header + 1, (ws.max_row or 0) + 1):
        id_kemenag = normalisasi.rapikan(_sel(ws, r, peta["id_kemenag"]))
        if not id_kemenag:
            continue
        ambil = lambda f: _sel(ws, r, peta[f]) if f in peta else None  # noqa: E731
        hasil.append({
            "id_kemenag": id_kemenag,
            "nama": normalisasi.rapikan(ambil("nama")),
            "kabupaten": normalisasi.kabupaten(ambil("kabupaten")),
            "kecamatan": normalisasi.kecamatan(ambil("kecamatan")),
            "desa": normalisasi.desa(ambil("desa")),
            "kategori": normalisasi.rapikan(ambil("kategori")),
            "tahun_berdiri": normalisasi.bilangan_bulat(ambil("tahun_berdiri")),
            "sumber": normalisasi.rapikan(ambil("sumber")),
            "keterangan": normalisasi.rapikan(ambil("keterangan")),
            "tipe_hak": normalisasi.rapikan(ambil("tipe_hak")),
            "status_sertipikat": normalisasi.rapikan(ambil("status_sertipikat")),
            "tahun": normalisasi.bilangan_bulat(ambil("tahun")),
            "nib": normalisasi.rapikan(ambil("nib")),
            "luas": normalisasi.angka(ambil("luas")),
            # Koordinat blok ini banyak yang rusak (511079.0, 0.0). Simpan apa adanya,
            # jangan pernah disalin ke objek_wakaf.
            "latitude": normalisasi.angka(ambil("latitude")),
            "longitude": normalisasi.angka(ambil("longitude")),
        })
    if hasil:
        laporan.kemenag_terbaca[ws.title] = len(hasil)
    return hasil


def baca_berkas(path) -> tuple[list[dict], list[dict], Laporan]:
    """Baca seluruh workbook. Tidak menyentuh database."""
    laporan = Laporan()
    wb = openpyxl.load_workbook(path, data_only=True)
    objek, kemenag = [], []
    for ws in wb.worksheets:
        if ws.title in SHEET_BUKAN_KECAMATAN:
            continue
        objek.extend(baca_sheet(ws, laporan))
        kemenag.extend(baca_kemenag(ws, laporan))
    wb.close()
    for baris in objek:
        kode = baris["tipologi_kode"] or "(kosong)"
        laporan.tipologi[kode] = laporan.tipologi.get(kode, 0) + 1
    return objek, kemenag, laporan
